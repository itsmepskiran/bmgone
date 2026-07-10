"""
Convert a CSV of agent/STP records into a formatted Excel workbook.

Meant to run after exporting a CSV from the "Image to Excel" admin
tool (payroll/image-to-excel.html) once its OCR results have been
manually reviewed/corrected there.

Expected CSV columns (header row required):
    RowID, AgentCode, FirstName, LastName, City, Pincode, Phone, Month

Usage:
    python csv_to_excel_agent_records.py input.csv output.xlsx

Requires: pandas, openpyxl
    pip install pandas openpyxl
"""

import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FORMULA_PREFIXES = ("=", "+", "-", "@")


def clean_text(value) -> str:
    """Blank for missing values; guards against Excel formula injection."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.startswith(FORMULA_PREFIXES):
        text = "'" + text
    return text


def build_workbook(csv_path: str, xlsx_path: str, sheet_name: str = "Agent STP Records"):
    # --- Load and de-duplicate ---
    df = pd.read_csv(csv_path, dtype=str)
    before = len(df)
    df = df.drop_duplicates(subset=["RowID", "Phone"])
    after = len(df)
    if before != after:
        print(f"Removed {before - after} duplicate row(s) (same RowID + Phone).")

    # --- Build workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = ["Row ID", "Agent Code", "First Name", "Last Name",
               "City", "Pincode", "Phone", "Month"]
    ws.append(headers)

    header_font = Font(bold=True, name="Arial", color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"  # keep header visible while scrolling

    body_font = Font(name="Arial")
    flag_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    flagged_count = 0

    for _, row in df.iterrows():
        row_id_raw = clean_text(row["RowID"])
        phone = clean_text(row["Phone"])
        pincode = clean_text(row["Pincode"])
        lastname = clean_text(row["LastName"])

        row_id_valid = row_id_raw.isdigit()
        row_id = int(row_id_raw) if row_id_valid else row_id_raw

        # Flag rows that look incomplete or malformed for manual review
        flagged = (
            not row_id_valid
            or len(phone) != 10
            or len(pincode) != 6
            or lastname == ""
        )
        if flagged:
            flagged_count += 1

        ws.append([
            row_id, clean_text(row["AgentCode"]), clean_text(row["FirstName"]), lastname,
            clean_text(row["City"]), pincode, phone, clean_text(row["Month"]),
        ])
        r = ws.max_row

        # Keep phone/pincode as text so leading digits/zeros are never dropped
        ws.cell(row=r, column=6).number_format = "@"  # Pincode
        ws.cell(row=r, column=7).number_format = "@"  # Phone

        for col in range(1, 9):
            ws.cell(row=r, column=col).font = body_font
            if flagged:
                ws.cell(row=r, column=col).fill = flag_fill

    # Column widths tuned for this data
    widths = [10, 11, 20, 16, 16, 10, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(xlsx_path)
    print(f"Saved {len(df)} rows to {xlsx_path} ({flagged_count} flagged for review)")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python csv_to_excel_agent_records.py input.csv output.xlsx")
        sys.exit(1)
    build_workbook(sys.argv[1], sys.argv[2])
